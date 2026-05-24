import MetaTrader5 as mt5
from openpyxl import load_workbook
import math

# =========================================
# EXCEL FILE
# =========================================

EXCEL_FILE = "risk_manage_file.xlsx"

# =========================================
# SHEET CONFIG
# =========================================

SHEET_NAME = "MASANIELLO"

START_ROW = 5

# =========================================
# GET ACTIVE ROW
# =========================================

def get_active_row():

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    row = START_ROW

    while True:

        result = ws[f"C{row}"].value

        # empty result means next trade
        if result is None or str(result).strip() == "":
            wb.close()
            return row

        row += 1

# =========================================
# GET RISK AMOUNT FROM COLUMN D
# =========================================

def get_risk_amount():

    row = get_active_row()

    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    risk_amount = ws[f"D{row}"].value

    wb.close()

    if risk_amount is None:
        return None

    return float(risk_amount)

# =========================================
# WRITE TRADE RESULT
# =========================================

def write_trade_result(result):

    # STRICTLY W or L
    result = str(result).upper()

    if result not in ["W", "L"]:
        return

    row = get_active_row()

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]

    ws[f"C{row}"] = result

    wb.save(EXCEL_FILE)
    wb.close()

# =========================================
# CALCULATE LOT SIZE
# =========================================

def calculate_lot_size(symbol, entry, sl):

    risk_amount = get_risk_amount()

    if risk_amount is None:
        return 0.01

    symbol_info = mt5.symbol_info(symbol)

    volume_min = symbol_info.volume_min
    volume_max = symbol_info.volume_max
    volume_step = symbol_info.volume_step

    order_type = mt5.ORDER_TYPE_BUY

    # start from minimum lot
    lot = volume_min

    best_lot = volume_min

    while lot <= volume_max:

        # MT5 real profit calculation
        profit = mt5.order_calc_profit(
            order_type,
            symbol,
            lot,
            entry,
            sl
        )

        if profit is None:
            break

        loss = abs(profit)

        # stay inside risk limit
        if loss <= risk_amount:
            best_lot = lot

        else:
            break

        lot += volume_step

    return round(best_lot, 2)

# =========================================
# CALCULATE BUY TP
# =========================================

def calculate_buy_tp(symbol, entry, sl, lot_size):

    risk_amount = get_risk_amount()

    reward_amount = risk_amount * 0.5

    symbol_info = mt5.symbol_info(symbol)

    point = symbol_info.point
    tick_value = symbol_info.trade_tick_value
    if tick_value is None or tick_value <= 0:
        tick_value = 1
    
    if lot_size <= 0:
        lot_size = 0.01
    money_per_point = tick_value * lot_size

    tp_points = reward_amount / money_per_point

    tp_price = entry + (tp_points * point)

    return round(tp_price, symbol_info.digits)

# =========================================
# CALCULATE SELL TP
# =========================================

def calculate_sell_tp(symbol, entry, sl, lot_size):

    risk_amount = get_risk_amount()

    reward_amount = risk_amount * 0.5

    symbol_info = mt5.symbol_info(symbol)

    point = symbol_info.point
    tick_value = symbol_info.trade_tick_value
    if tick_value is None or tick_value <= 0:
        tick_value = 1
    
    if lot_size <= 0:
        lot_size = 0.01
    money_per_point = tick_value * lot_size

    tp_points = reward_amount / money_per_point

    tp_price = entry - (tp_points * point)

    return round(tp_price, symbol_info.digits)

# =========================================
# CHECK CLOSED TRADE RESULT
# =========================================

def check_trade_result(position_ticket):

    deals = mt5.history_deals_get(position=position_ticket)

    if deals is None:
        return None

    total_profit = 0

    for deal in deals:
        total_profit += deal.profit

    if total_profit > 0:
        return "W"

    elif total_profit < 0:
        return "L"

    return None

# =========================================
# PROCESS CLOSED TRADE
# =========================================

def process_closed_trade(result):

    result = str(result).upper()

    if result not in ["W", "L"]:
        return

    write_trade_result(result)

    print(f"RESULT WRITTEN TO EXCEL: {result}")